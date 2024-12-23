from flask import Flask, request, jsonify
from flask_cors import CORS
from transformers import pipeline
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)  # Enables cross-origin requests

# Load emotion detection model
emotion_classifier = pipeline("text-classification", model="j-hartmann/emotion-english-distilroberta-base", return_all_scores=True)

@app.route('/upload-excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    workbook = load_workbook(file)
    sheet = workbook.active

    emotions_summary = []

    # Process each row in the Excel sheet
    for row in sheet.iter_rows(values_only=True):
        comment = row[0]  # Assuming comments are in the first column
        if comment:
            emotions = emotion_classifier(comment)
            emotions_summary.append({
                "comment": comment,
                "emotions": {e['label']: e['score'] for e in emotions[0]}
            })

    return jsonify(emotions_summary)

@app.route('/analyze-text', methods=['POST'])
def analyze_text():
    data = request.json
    text = data.get('text', '')

    if not text:
        return jsonify({"error": "No text provided"}), 400

    emotions = emotion_classifier(text)
    return jsonify({
        "text": text,
        "emotions": {e['label']: e['score'] for e in emotions[0]}
    })

if __name__ == '__main__':
    app.run(debug=True)
